# utils/excel_utils.py (With new detailed ledger export function)
import discord
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import io
import logging
from datetime import datetime
import database

logger = logging.getLogger('discord_bot_excel_utils')

def apply_header_style(worksheet, row, columns):
    """Apply formatting to header row cells."""
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col in range(1, columns + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        worksheet.column_dimensions[get_column_letter(col)].width = 20

def auto_adjust_column_widths(worksheet):
    """Auto-adjust all column widths on a worksheet based on content."""
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = min(adjusted_width, 60)

async def create_excel_export(guild_id: int, guild_name: str, bot: discord.Client) -> io.BytesIO | None:
    logger.info(f"Starting Excel export for guild: {guild_name} ({guild_id})")
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    guild = bot.get_guild(guild_id)
    if not guild: return None

    # Sheet 1: User Balances
    ws_users = workbook.create_sheet(title="User Balances")
    headers_users = ["User ID", "User Name", "Balance"]
    ws_users.append(headers_users)
    apply_header_style(ws_users, 1, len(headers_users))
    users_data = await database.get_all_users(guild_id, guild_name)
    for user in users_data: ws_users.append([user.get('user_id'), user.get('user_name'), user.get('balance')])
    auto_adjust_column_widths(ws_users)

    # Sheet 2: All Transactions
    ws_all_trans = workbook.create_sheet(title="All Transactions")
    headers_all_trans = ["Trans ID", "User ID", "User Name", "Type", "Amount", "Reason", "Timestamp (UTC)", "Task Name", "Channel ID", "Forum ID", "Post ID", "Performed By Name"]
    ws_all_trans.append(headers_all_trans)
    apply_header_style(ws_all_trans, 1, len(headers_all_trans))
    all_transactions_data = await database.get_all_transactions(guild_id, guild_name)
    for trans in all_transactions_data: ws_all_trans.append([trans.get('transaction_id'), trans.get('user_id'), trans.get('user_name'), trans.get('type'), trans.get('amount'), trans.get('reason'), trans.get('timestamp'), trans.get('task_name'), trans.get('channel_id'), trans.get('forum_id'), trans.get('post_id'), trans.get('added_by_name')])
    auto_adjust_column_widths(ws_all_trans)
    
    # Sheet 3: Tasks
    ws_tasks = workbook.create_sheet(title="Tasks")
    headers_tasks = ["Task Name", "Rank"]
    ws_tasks.append(headers_tasks)
    apply_header_style(ws_tasks, 1, len(headers_tasks))
    tasks_data = await database.get_tasks(guild_id, guild_name)
    for task in tasks_data: ws_tasks.append([task.get('task_name'), task.get('rank')])
    auto_adjust_column_widths(ws_tasks)
    
    # Sheet 4: Default Rates
    ws_default_rates = workbook.create_sheet(title="Default Rates")
    headers_default_rates = ["Task Name", "Rate"]
    ws_default_rates.append(headers_default_rates)
    apply_header_style(ws_default_rates, 1, len(headers_default_rates))
    default_rates_data = await database.get_default_rates(guild_id, guild_name)
    for task, rate in default_rates_data.items(): ws_default_rates.append([task, rate])
    auto_adjust_column_widths(ws_default_rates)

    # Sheet 5: Channel Rates
    ws_channel_rates = workbook.create_sheet(title="Channel Rates")
    headers_channel_rates = ["Channel ID", "Channel Name", "Task Name", "Rate"]
    ws_channel_rates.append(headers_channel_rates)
    apply_header_style(ws_channel_rates, 1, len(headers_channel_rates))
    for channel in guild.text_channels:
        channel_rates = await database.get_channel_rates(guild_id, guild_name, channel.id)
        for task, rate in channel_rates.items(): ws_channel_rates.append([channel.id, channel.name, task, rate])
    auto_adjust_column_widths(ws_channel_rates)

    # Sheet 6: Forum Rates
    ws_forum_rates = workbook.create_sheet(title="Forum Rates")
    headers_forum_rates = ["Forum ID", "Forum Name", "Task Name", "Rate"]
    ws_forum_rates.append(headers_forum_rates)
    apply_header_style(ws_forum_rates, 1, len(headers_forum_rates))
    for forum in guild.forums:
        forum_rates = await database.get_forum_rates(guild_id, guild_name, forum.id)
        for task, rate in forum_rates.items(): ws_forum_rates.append([forum.id, forum.name, task, rate])
    auto_adjust_column_widths(ws_forum_rates)

    # Sheet 7: Post Rates
    ws_post_rates = workbook.create_sheet(title="Post Rates")
    headers_post_rates = ["Post ID", "Post Name", "Parent Forum", "Task Name", "Rate"]
    ws_post_rates.append(headers_post_rates)
    apply_header_style(ws_post_rates, 1, len(headers_post_rates))
    all_threads = list(guild.threads)
    for forum in guild.forums:
        try:
            async for thread in forum.archived_threads(limit=None): all_threads.append(thread)
        except discord.Forbidden: continue
    for thread in all_threads:
        post_rates = await database.get_post_rates(guild_id, guild_name, thread.id)
        for task, rate in post_rates.items(): ws_post_rates.append([thread.id, thread.name, thread.parent.name, task, rate])
    auto_adjust_column_widths(ws_post_rates)
    
    # Sheet 8: Contact Info
    ws_contact = workbook.create_sheet(title="Contact Info")
    headers_contact = ["User ID", "User Name", "Email", "Payment Method"]
    ws_contact.append(headers_contact)
    apply_header_style(ws_contact, 1, len(headers_contact))
    contact_info_data = await database.get_all_contact_info_for_export(guild_id, guild_name)
    for contact in contact_info_data: ws_contact.append([contact.get('user_id'), contact.get('user_name'), contact.get('email'), contact.get('payment_method')])
    auto_adjust_column_widths(ws_contact)
    
    excel_data = io.BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)
    return excel_data

async def create_user_ledgers_export(guild_id: int, guild_name: str) -> io.BytesIO | None:
    logger.info(f"Starting DETAILED user ledger export for guild: {guild_name} ({guild_id})")
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    all_users = await database.get_all_users(guild_id, guild_name)
    if not all_users:
        ws = workbook.create_sheet(title="No Users Found")
        ws['A1'] = "No users with balances were found in the database."
    
    sorted_users = sorted(all_users, key=lambda u: u.get('user_name', '').lower())

    for user in sorted_users:
        user_id = user.get('user_id')
        user_name = user.get('user_name', str(user_id))
        sheet_title = database.sanitize_name(user_name)[:31]
        ws = workbook.create_sheet(title=sheet_title)

        headers = ["Trans ID", "Timestamp (UTC)", "Type", "Amount", "Reason", "Task Name", "Context ID", "Logged By"]
        ws.append(headers)
        apply_header_style(ws, 1, len(headers))

        user_transactions = await database.get_user_transactions(guild_id, guild_name, user_id, limit=None)
        for trans in user_transactions:
            context_id = trans.get('post_id') or trans.get('forum_id') or trans.get('channel_id') or 'N/A'
            ws.append([trans.get('transaction_id'), trans.get('timestamp'), trans.get('type', 'N/A').capitalize(), trans.get('amount'), trans.get('reason'), trans.get('task_name'), context_id, trans.get('added_by_name')])
        auto_adjust_column_widths(ws)

    excel_data = io.BytesIO()
    workbook.save(excel_data)
    excel_data.seek(0)
    return excel_data